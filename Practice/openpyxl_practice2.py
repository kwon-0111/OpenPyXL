from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()

dest_filename = 'empty_book.xlsx'

ws1 = wb.active
ws1.title = "range names"

# 39갱의 행에 0부터 599까지의 값 추가
for row in range(1, 40):
    ws1.append(range(600))

ws2 = wb.create_sheet(title = "Pi")
ws2["F5"] = 3.14

ws3 = wb.create_sheet(title = "Data")

# 각 셀에 해당 셀의 열 번호를 문자로 입력
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value=get_column_letter(col))

# 각 셀에 해당 셀의 열 번호를 문자로 입력
# S.format() 메소드
    # S: 셀에 입력할 문자열, 서식지정자와 비슷하게 {정수}사용 가능
    # format의 parameter: {정수}에 넣을 값들
for row in range(21, 30):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}: {1}".format("열", get_column_letter(col)))

# 이미 존재하는 워크북 읽어들이기   
from openpyxl import load_workbook
# load_workbook의 flag(data_only 등) 사용 가능
# openpyxl does currently not read all possible items in an Excel file
# so images and charts will be lost from existing files if they are opened and saved with the same name
wb2 = load_workbook(filename='play.xlsx')
print(wb2['Sheet1']['A1'].value)
print('--------------------')

# Using number formats
import datetime

ws3['A1'] = datetime.datetime(2022, 10, 23)
print(ws3['A1'].number_format)  

# Using formulae
# 수식을 문자열 취급하여 사용하면 된다
ws3['B1'] = ws3['B2'] = 10
ws3['B3'] = '=SUM(B1:B2)'

# Merge / Unmerge cells
# 병합된 셀의 데이터는 사라진다
ws3.merge_cells('B2:B3')
ws3.unmerge_cells('B2:B3')

# Fold
ws3.column_dimensions.group('A','D', hidden=True)
ws3.row_dimensions.group(1,10, hidden=True)

wb.save(dest_filename)
