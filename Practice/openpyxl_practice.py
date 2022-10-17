# OpenPyXL 공부
# official documentation 보고 공부함
# 2022.10.16

# Workbook 클래스 import 하기
from openpyxl import Workbook

# 워크북(엑셀 파일) 생성
wb = Workbook()

# active : Get the currently active sheet or None, 값 변경 하지 않는 한 첫번째 시트
ws1 = wb.active

# ws1 시트의 이름 변경
ws1.title = "ws1"

# 맨 뒤에 새로운 시트 추가
ws2 = wb.create_sheet("ws2")

# 2번째 위치에 새로운 시트 추가
ws3 = wb.create_sheet("ws3", 2)

# 맨 앞에 새로운 시트 추가
ws0 = wb.create_sheet("ws0", 0)

# change the background color of the tab 
ws0.sheet_properties.tabColor = "FF0000"

# 시트 접근
subName = wb["ws3"]
subName.sheet_properties.tabColor = "FFFF00"

# Workbook.sheetname 속성을 이용하여 모든 시트 이름 열람 가능
print(wb.sheetnames)
print("------------------------------")

# for문을 이용한 시트 접근
for sheet in wb: 
    print(sheet.sheet_properties.tabColor, '\n')
print("------------------------------")

# "같은 워크북 내에서" 워크시트 복사하기(다른 워크북끼리는 불가, 셀의 값 등만 복사, 차트/그림 등은 복사 불가)
source = wb.active
target = wb.copy_worksheet(source)

# 하나의 셀에 접근
# !! 셀에 처음 접근할 때 워크시트에 해당 셀이 생성된다 !!
ws0["A4"] = 2
cell_val1 = ws0["A4"]
print(cell_val1.value)

cell_val2 = ws0.cell(row = 4, column = 1, value = 3)
print(cell_val2.value)
print("------------------------------")

# 여러 셀에 접근
# 셀에 처음 접근할 때 해당 셀이 생성되기 때문에 값은 비어 있어도 해당 범위의 셀들은 생성되어 있다
cell_range1 = ws0["A1" : "C2"]
colC = ws0["C"]
col_range = ws0["C" : "D"]
row2 = ws0[10]
row_range = ws0[5:10]

# iterator를 이용한 접근(performance 때문에 read only 모드에서는 사용 불가  )
for row in ws0.iter_rows(min_row=1, min_col=3, max_row=2, max_col=4): # row iterator
    for cell in row:
        print(cell)
print("------------------------------")

for row in ws0.iter_cols(min_row=1, min_col=3, max_row=2, max_col=4): # col iterator
    for cell in row:
        print(cell)
print("------------------------------")

# 시트 내 모든 row 또는 column을 iterate 해야한다면 
# 위에서 row를 10까지 접근했기 때문에 10까지 iterate한다
print(tuple(ws0.rows))
print("------------------------------")
print(tuple(ws0.columns))
print("------------------------------")

# 모든 row를 iterate하지만 셀 값만 반환
# 어차피 모든 data를 iterate할 거라 row로만 가능
for row in ws0.values:
    for value in row:
        print(value)
print("------------------------------")

# Worksheet.iter_rows()와 Worksheet.iter_cols() 모두 values_only parameter를 받아들여 셀의 값만 반환 가능
for row in ws0.iter_rows(min_row=3, min_col=1, max_row=4, max_col=2, values_only=True):
    for cell in row:
        print(cell)
print("------------------------------")

# 파일(워크북) 저장
# !!덮어쓴다!!
wb.save("prac.xlsx")

# 파일(워크북) 불러오기
from openpyxl import load_workbook
wb2 = load_workbook("prac.xlsx")
print(wb2.sheetnames)
