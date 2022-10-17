from openpyxl import Workbook
import random

wb = Workbook()

Sheet1 = wb.active
Sheet1.title = "Sheet1"

def sort_one_row(row_number):
    arr = []
    for row in Sheet1.iter_rows(min_row  = row_number, max_row = row_number, min_col = 1, max_col = 10):
        for cell in row:
            arr.append(cell.value)
    arr.sort(reverse=True)

    for row in Sheet1.iter_rows(min_row = row_number, max_row = row_number, min_col = 1, max_col = 10):
        for cell in row:
            cell.value = arr.pop()
            # print(cell, end = ' ')
        

for row in Sheet1.iter_rows(min_row = 1, max_row = 10, min_col = 1, max_col = 10):
    for cell in row:
        cell.value = random.randint(1,100)

for i in range(1,11):
    sort_one_row(i)

print('\n')


wb.save("play.xlsx")
