from os import remove
import openpyxl
from copy import copy

from openpyxl import Workbook
from openpyxl import load_workbook

import datetime

min_row_of_monthly_total = 4
min_row_of_monthly_total = 53

def copy_sheet(source_sheet, target_sheet):
    copy_cells(source_sheet, target_sheet)  # copy all the cell values and styles
    copy_sheet_attributes(source_sheet, target_sheet)
def copy_sheet_attributes(source_sheet, target_sheet):
    if isinstance(source_sheet, openpyxl.worksheet._read_only.ReadOnlyWorksheet):
        return
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.merged_cells = copy(source_sheet.merged_cells)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.freeze_panes = copy(source_sheet.freeze_panes)

    # set row dimensions
    # So you cannot copy the row_dimensions attribute. Does not work (because of meta data in the attribute I think). So we copy every row's row_dimensions. That seems to work.
    for rn in range(len(source_sheet.row_dimensions)):
        target_sheet.row_dimensions[rn] = copy(source_sheet.row_dimensions[rn])

    if source_sheet.sheet_format.defaultColWidth is None:
        print('Unable to copy default column wide')
    else:
        target_sheet.sheet_format.defaultColWidth = copy(source_sheet.sheet_format.defaultColWidth)

    # set specific column width and hidden property
    # we cannot copy the entire column_dimensions attribute so we copy selected attributes
    for key, value in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[key].min = copy(source_sheet.column_dimensions[key].min)   # Excel actually groups multiple columns under 1 key. Use the min max attribute to also group the columns in the targetSheet
        target_sheet.column_dimensions[key].max = copy(source_sheet.column_dimensions[key].max)  # https://stackoverflow.com/questions/36417278/openpyxl-can-not-read-consecutive-hidden-columns discussed the issue. Note that this is also the case for the width, not onl;y the hidden property
        target_sheet.column_dimensions[key].width = copy(source_sheet.column_dimensions[key].width) # set width for every column
        target_sheet.column_dimensions[key].hidden = copy(source_sheet.column_dimensions[key].hidden)
def copy_cells(source_sheet, target_sheet):
    for r, row in enumerate(source_sheet.iter_rows()):
        for c, cell in enumerate(row):
            source_cell = cell
            if isinstance(source_cell, openpyxl.cell.read_only.EmptyCell):
                continue
            target_cell = target_sheet.cell(column=c+1, row=r+1)

            target_cell._value = source_cell._value
            target_cell.data_type = source_cell.data_type

            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

            if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)

            if not isinstance(source_cell, openpyxl.cell.ReadOnlyCell) and source_cell.comment:
                target_cell.comment = copy(source_cell.comment)

def input_file():
    input_quotation = str(input("Enter the name and path of the quotation file: "))
    wb_input_quotation = load_workbook(filename='C:\\Users\\Alex\\Desktop\\Project_22_10_23\\' + input_quotation)
    return wb_input_quotation

# 워크시트에 input_val 값이 있는지 확인 후 해당 셀 반환
def search_cell_by_value(input_ws, input_val):
    for row in input_ws.iter_rows(min_row=1, min_col=1, max_row=60, max_col=12):
        for cell in row:
            if cell.value == input_val:
                return cell
    print('찾는 값이 해당 워크시트에 존재하지 않습니다')

# def fill_in_DATE():
#     _ = search_cell_by_value(wb_input_quotation['풀빌라'], '기       간')
#     cell_with_data = wb_input_quotation['풀빌라'].cell(row = _.row, column = _.column+2)
#     value = cell_with_data.value
#     year_month = value.split('.')[0] + '.' + value.split('.')[1]

#     if not is_worksheet_in_workbook(wb_monthly_total, year_month):
#         wb = wb_monthly_total
#         wb_source = load_workbook('monthly_total_form.xlsx', data_only=True, read_only=True)
#         source_sheet = wb_source['Sheet1']
#         ws = wb.create_sheet(year_month)
#         copy_sheet(source_sheet, ws)

#     wb_monthly_total[year_month]['B5'] = '20' + value.split('.')[0] + '-' + value.split('.')[1] + '-' + value.split('.')[2]   

# datetime.date 형식으로 반환  
def get_date():
    _ = search_cell_by_value(wb_input_quotation['풀빌라'], '기       간')
    cell_with_data = wb_input_quotation['풀빌라'].cell(row = _.row, column = _.column+2)
    value = cell_with_data.value
    value = datetime.date(int('20' + value.split('.')[0]), int(value.split('.')[1]), int(value.split('.')[2]))
    return value

# 23.01 형식으로 반환
def get_year_month(input_date):
    _ = str(input_date)
    year_month = str(int(_.split('-')[0]) % 100) + '.' + _.split('-')[1]
    return year_month

# input_date의 년도와 달을 이름으로 가진 워크시트가 있는지 체크
def create_worksheet_if_not_in_workbook(input_date):
    year_month = get_year_month(input_date)

    for sheet in wb_monthly_total.sheetnames:
        if sheet == year_month:
            return
    
    wb = wb_monthly_total
    wb_source = load_workbook('monthly_total_form.xlsx', data_only=True, read_only=True)
    source_sheet = wb_source['Sheet1']
    ws = wb.create_sheet(get_year_month(input_date))
    copy_sheet(source_sheet, ws)




def fill_in_NAME():
    _ = search_cell_by_value(wb_input_quotation['풀빌라'], ' 여  행  명')
    cell_with_data = wb_input_quotation['풀빌라'].cell(row = _.row, column = _.column+2)
    value = cell_with_data.value.split('님 ')[0]
    

# def fill_in_NIGHT():

# def fill_in_PAX(): 

# def fill_in_AMOUNT_TOTAL(): 

# def fill_in_AMOUNT_PER_HEAD():



if __name__ == '__main__':

    wb_monthly_total = load_workbook(filename='C:\\Users\\Alex\\Desktop\\Project_22_10_23\\monthly_total.xlsx')

    # # 복사 후, Sheet라'는 이름을 가진 워크시트가 남아있으면 삭제
    # if 'Sheet' in wb_monthly_total.sheetnames:
    #     wb_monthly_total.remove(wb_monthly_total['Sheet'])

    #  wb_input_quotation = input_file()
    wb_input_quotation = load_workbook(filename='C:\\Users\\Alex\\Desktop\\Project_22_10_23\\ex_quotation.xlsx')
    
    # fill_in_DATE()
    # fill_in_NAME()
   
    input_date = get_date()
    create_worksheet_if_not_in_workbook(input_date)
    
    

    wb_monthly_total.save('monthly_total.xlsx')
